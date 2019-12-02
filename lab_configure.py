
#_*_coding=utf-8_*_

from openpyxl import load_workbook
import re

def generate_con_sheet(equipment_list,wb):
	for i in equipment_list:
		wb.create_sheet(i)
	print('generate configure store sheets as blew ' + str(wb.sheetnames))

def get_equipment_number(topology_sheet):
	#打开文档
	n = 20 #本excel最多支持16台设备,其实只能支持10个
	equipment_dist       = {} #定义一个空字典
	equipment_count_dist = {}
	equipment_list       = []
	for n in range(3,20):

		equipment_name   = topology_sheet.cell(n,1).value
		equipment_number = topology_sheet.cell(n,2).value
		inital_number    = 0
		#如果是空值无法加入字典的
		if equipment_name == None:
			pass
		else:
			equipment_dist.update({equipment_name:equipment_number})
			#向字典中添加元素
			equipment_count_dist.update({equipment_name:inital_number})
			#生成各个设备出现频率计数用的字典
			equipment_list.append(equipment_name)

		n = n + 1
	print('生成的设备列表如下：')
	print(equipment_dist)
	return equipment_dist

def get_equipment_count_dist(topology_sheet):
	#打开文档
	n = 20 #本excel最多支持16台设备,其实只能支持10个
	equipment_dist       = {} #定义一个空字典
	equipment_count_dist = {}
	equipment_list       = []
	for n in range(3,20):

		equipment_name   = topology_sheet.cell(n,1).value
		equipment_number = topology_sheet.cell(n,2).value
		inital_number    = 0
		#如果是空值无法加入字典的
		if equipment_name == None:
			pass
		else:
			equipment_dist.update({equipment_name:equipment_number})
			#向字典中添加元素
			equipment_count_dist.update({equipment_name:inital_number})
			#生成各个设备出现频率计数用的字典
			equipment_list.append(equipment_name)

		n = n + 1
	print('生成的设备初始字典如下： ')
	print(equipment_count_dist)
	return equipment_count_dist

def get_equipment_line_dist(topology_sheet):
	#打开文档
	n = 20 #本excel最多支持16台设备,其实只能支持10个
	equipment_dist       = {} #定义一个空字典
	equipment_count_dist = {}
	equipment_list       = []
	equipment_line_dist  = {}
	for n in range(3,20):

		equipment_name   = topology_sheet.cell(n,1).value
		equipment_number = topology_sheet.cell(n,2).value
		inital_number    = 0
		line_number      = 1
		#如果是空值无法加入字典的
		if equipment_name == None:
			pass
		else:
			equipment_dist.update({equipment_name:equipment_number})
			#向字典中添加元素
			equipment_count_dist.update({equipment_name:inital_number})
			#生成各个设备出现频率计数用的字典
			equipment_list.append(equipment_name)
			equipment_line_dist.update({equipment_name:line_number})

		n = n + 1
	print('生成的设备命令行字典如下： ')
	print(equipment_line_dist)
	return equipment_line_dist

def get_equipment_list(topology_sheet):
	#打开文档
	n = 20 #本excel最多支持16台设备,其实只能支持10个
	equipment_dist       = {} #定义一个空字典
	equipment_count_dist = {}
	equipment_list       = []
	for n in range(3,20):

		equipment_name   = topology_sheet.cell(n,1).value
		equipment_number = topology_sheet.cell(n,2).value
		inital_number    = 0
		#如果是空值无法加入字典的
		if equipment_name == None:
			pass
		else:
			equipment_dist.update({equipment_name:equipment_number})
			#向字典中添加元素
			equipment_count_dist.update({equipment_name:inital_number})
			#生成各个设备出现频率计数用的字典
			equipment_list.append(equipment_name)

		n = n + 1
	print('生成的设备名称列表为： ')
	print(equipment_list)
	return equipment_list

def generate_ip_left(topology_sheet,equipment_dist,n):
	#子模块不能做循环，除非要输出字典和列表。
	#如果需要循环读取，可以在main模块里面循环调用

	#n由上级函数传递

	equipment_name  = topology_sheet.cell(n,5).value
	remot_equipment = topology_sheet.cell(n,6).value
	#print(equipment_name)
	link_left  = equipment_dist[equipment_name]
	link_right = equipment_dist[remot_equipment]

	subnet  = str(link_left) + str(link_right)

	ip_left  = subnet + '.0.0.' + str(link_left)
	ip_right = subnet + '.0.0.' + str(link_right)
	return ip_left

def generate_ip_right(topology_sheet,equipment_dist,n):
	#子模块不能做循环，除非要输出字典和列表。
	#如果需要循环读取，可以在main模块里面循环调用

	equipment_name  = topology_sheet.cell(n,5).value
	remot_equipment = topology_sheet.cell(n,6).value

	link_left  = equipment_dist[equipment_name]
	link_right = equipment_dist[remot_equipment]

	subnet  = str(link_left) + str(link_right)

	ip_left  = subnet + '.0.0.' + str(link_left)
	ip_right = subnet + '.0.0.' + str(link_right)

	return ip_right

def caculate_interface(topology_sheet):

	n = 1
	intf_list = [] #这种变量不能再循环里面定义，会被重置

	for n in range(47,54):
		intf_type      =  topology_sheet.cell(n,1).value
		if intf_type == None:
			pass
		else:
			intf_slot       =  topology_sheet.cell(n,2).value
			intf_number     =  topology_sheet.cell(n,3).value
			intf_start_num  = topology_sheet.cell(n,4).value
			intf_end_num    = int(intf_number) - int(intf_start_num) - 1

			while int(intf_start_num) <= int(intf_end_num):
				intf_name   = str(intf_type) + str(intf_slot) + '/' + str(intf_start_num)
				intf_list.append(str(intf_name))
				intf_start_num = intf_start_num + 1
			else:
				pass
	print('生成接口的列表如下： ')
	print(intf_list)
	return intf_list

def write_cable_list(cable_list_sheet,topology_sheet,equipment_dist,equipment_count_dist,intf_list):
	#def 可以传递进很多的参数，但是只能return一个数值；但是，它可以对外部对象做操作。
	#1.下级函数是计算功能的，需要直接调用运算
	#2.下级函数仅仅是用于计算出一个结果的，可以直接将结果作为参数输入

	n = 1#这里无所谓，都是从for语句来的
	i = 2#cable-list行数，从第二行开始
	equipment_count_dist_local = equipment_count_dist
	for n in range(3,30):

		if topology_sheet.cell(n,5).value == None:
			pass
		else:
			cable_list_sheet.cell(i,1).value  = topology_sheet.cell(n,5).value
			cable_list_sheet.cell(i,11).value = topology_sheet.cell(n,6).value

			cable_list_sheet.cell(i,5).value = generate_ip_left(topology_sheet,equipment_dist,n)
			cable_list_sheet.cell(i,7).value = generate_ip_right(topology_sheet,equipment_dist,n)

			cable_list_sheet.cell(i,6).value = '255.255.255.0'
			cable_list_sheet.cell(i,8).value = '255.255.255.0'
			#MPLS、OSPF、OSPF AREA
			cable_list_sheet.cell(i,13).value = topology_sheet.cell(n,7).value
			cable_list_sheet.cell(i,14).value = topology_sheet.cell(n,8).value
			cable_list_sheet.cell(i,15).value = '0'

			equipment_name_left          =  cable_list_sheet.cell(i,1).value
			#print('left')
			#print(equipment_name_left)
			equipment_intf_number        =  equipment_count_dist[equipment_name_left]
			#print(equipment_intf_number)
			intf_left                    =  intf_list[equipment_intf_number]
			#第几次使用就读取第几个接口
			equipment_intf_number_update =  int(equipment_intf_number) + 1
			#print(equipment_intf_number_update)
			equipment_count_dist.update({equipment_name_left:equipment_intf_number_update})
			#print(equipment_count_dist)

			#遇到一次就加一，说明这个对应接口已经使用

			equipment_name_right         =  cable_list_sheet.cell(i,11).value
			#print('right')
			#print(equipment_name_right)
			equipment_intf_number        =  equipment_count_dist[equipment_name_right]
			#print(equipment_intf_number)
			intf_right                   =  intf_list[equipment_intf_number]
			equipment_intf_number_update =  int(equipment_intf_number) + 1 #如果使用了这个接口计数器加一
			#print(equipment_intf_number_update)
			equipment_count_dist.update({equipment_name_right:equipment_intf_number_update})
			#print(equipment_count_dist)
			cable_list_sheet.cell(i,2).value  = intf_left
			cable_list_sheet.cell(i,9).value  = intf_right

			i = i + 1#n会自动循环

	print('已经完成')

def configure_by_cable_list(wb,equipment_line_dist):

	#configuration

	topology_sheet = wb['topology']
	cable_list_sheet = wb['cable_list']


	n = 1 #cable-list的行数
	i = 1 #Loopback行数
	L = 1#从第一行开始

	for i in range(3,20):
		equipment_name      = topology_sheet.cell(i,1).value
		loopback_ip         = topology_sheet.cell(i,3).value
		if equipment_name   == None:
			pass
		else:
			sheet_equipment  = wb[equipment_name]
			L = equipment_line_dist[equipment_name]

			sheet_equipment.cell(L,1).value   = 'interface loopback 1'
			L = L + 1
			sheet_equipment.cell(L,1).value   = 'ip address ' + loopback_ip
			L = L + 1
			sheet_equipment.cell(L,1).value   = '!'
			L = L + 1
		equipment_line_dist.update({equipment_name:L})

	print('Loopback IP的配置已经输出完毕！！！')

	for n in range(2,30):

		equipment_left    = cable_list_sheet.cell(n,1).value
		if equipment_left == None:
			pass
		else:
			intf_left       = cable_list_sheet.cell(n,2).value
			ip_left         = cable_list_sheet.cell(n,5).value
			mask_left       = cable_list_sheet.cell(n,6).value
			mpls            = cable_list_sheet.cell(n,13).value
			ospf_enable     = cable_list_sheet.cell(n,14).value
			ospf_area       = cable_list_sheet.cell(n,15).value

			equipment_right = cable_list_sheet.cell(n,11).value
			ip_right        = cable_list_sheet.cell(n,7).value
			mask_right      = cable_list_sheet.cell(n,8).value
			intf_right      = cable_list_sheet.cell(n,9).value

			sheet_left      = wb[equipment_left]
			L = equipment_line_dist[equipment_left]

			sheet_left.cell(L,1).value     =  'interface ' + intf_left
			L = L + 1
			sheet_left.cell(L,1).value     =  'ip address ' + ip_left + ' ' + mask_left
			L = L + 1
			sheet_left.cell(L,1).value     =   '!'
			L = L + 1
			if mpls == 'YES':

				sheet_left.cell(L,1).value     =  'interface ' + intf_left
				L = L + 1
				sheet_left.cell(L,1).value     =  'mpls ip'
				L = L + 1
				sheet_left.cell(L,1).value     =   '!'
				L = L + 1
			else:
				pass

			if ospf_enable == 'YES':
				sheet_left.cell(L,1).value     =   'router ospf 1'
				L = L + 1
				sheet_left.cell(L,1).value     =    'network ' + ip_left + ' ' + mask_left + ' are ' + ospf_area
				L = L + 1
				sheet_left.cell(L,1).value     =    '!'
				L = L + 1
			else:
				pass
			equipment_line_dist[equipment_left] = L #更新L的数值

			sheet_right      = wb[equipment_right]
			L = equipment_line_dist[equipment_right]

			sheet_right.cell(L,1).value     =  'interface ' + intf_right
			L = L + 1
			sheet_right.cell(L,1).value     =  'ip address ' + ip_right + ' ' + mask_right
			L = L + 1
			sheet_right.cell(L,1).value     =   '!'
			L = L + 1
			if mpls == 'YES':

				sheet_right.cell(L,1).value     =  'interface ' + intf_right
				L = L + 1
				sheet_right.cell(L,1).value     =  'mpls ip'
				L = L + 1
				sheet_right.cell(L,1).value     =   '!'
				L = L + 1
			else:
				pass

			if ospf_enable == 'YES':
				sheet_right.cell(L,1).value     =   'router ospf 1'
				L = L + 1
				sheet_right.cell(L,1).value     =    'network ' + ip_right + ' ' + mask_right + ' are ' + ospf_area
				L = L + 1
				sheet_right.cell(L,1).value     =    '!'
				L = L + 1
			else:
				pass
			equipment_line_dist[equipment_right] = L #更新L的数值

	print('已经完成配置输出！！！')


def main():

	#打开Excel表，打开特定的sheet
	Excel_Name             = 'topology.xlsx'
	wb                     = load_workbook(Excel_Name)
	topology_sheet         = wb['topology']
	cable_list_sheet       = wb['cable_list']

	#
	equipment_dist         = get_equipment_number(topology_sheet)
	equipment_count_dist   = get_equipment_count_dist(topology_sheet)
	equipment_list         = get_equipment_list(topology_sheet)
	equipment_line_dist    = get_equipment_line_dist(topology_sheet)
	#ip_left                = generate_ip_left(topology_sheet,equipment_dist)
	#ip_right               = generate_ip_right(topology_sheet,equipment_dist)
	intf_list              = caculate_interface(topology_sheet)
	write_cable_list(cable_list_sheet,topology_sheet,equipment_dist,equipment_count_dist,intf_list)

	#添加各个设备的sheet，用于保存配置
	generate_con_sheet(equipment_list,wb)
	#输出配置
	configure_by_cable_list(wb,equipment_line_dist)
	#保存文档
	wb.save('configure_topology20191129.xlsx')

if __name__ == '__main__':
    main()

    #con_wb.save(configure_output_Excel_Name)







